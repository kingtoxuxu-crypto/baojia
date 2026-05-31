from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "卫浴陶瓷报价表_第一版改造_20260529.xlsx"


def quote_sheet_name(name: str) -> str:
    return f"'{name}'" if any(ch in name for ch in " -") else name


def add_name(wb: Workbook, name: str, sheet: str, ref: str) -> None:
    wb.defined_names.add(DefinedName(name, attr_text=f"{quote_sheet_name(sheet)}!{ref}"))


def style_range(ws, cell_range: str, fill=None, font=None, alignment=None, border=None) -> None:
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_list_validation(ws, formula: str, ranges: list[str], allow_blank: bool = False) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = "请选择下拉列表中的标准选项。"
    dv.errorTitle = "选项不匹配"
    ws.add_data_validation(dv)
    for rng in ranges:
        dv.add(rng)


def build() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "填写首页"
    pack_info = wb.create_sheet("产品包装信息")
    load_calc = wb.create_sheet("装柜测算")
    check = wb.create_sheet("价格口径检查")
    bom = wb.create_sheet("BOM")
    log_params = wb.create_sheet("物流参数")
    log_calc = wb.create_sheet("物流测算")
    customer = wb.create_sheet("CustomerQuote")
    internal = wb.create_sheet("InternalCost")
    lists = wb.create_sheet("Lists")

    # Palette
    navy = "1F4E78"
    blue = "D9EAF7"
    light_yellow = "FFF2CC"
    light_green = "E2F0D9"
    light_orange = "FCE4D6"
    gray = "E7E6E6"
    dark_gray = "666666"
    white = "FFFFFF"
    red = "F4CCCC"
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    title_font = Font(name="Microsoft YaHei", bold=True, size=16, color=white)
    header_font = Font(name="Microsoft YaHei", bold=True, color=white)
    normal_font = Font(name="Microsoft YaHei", size=10)

    # Lists and named ranges
    list_data = {
        "A": ["尼厂出厂成本", "尼厂装柜成本", "FOB尼日利亚港", "手动确认成本"],
        "B": ["EXW中国供应商", "送中国指定仓/进仓价", "FOB中国港", "到尼厂入库成本", "手动确认成本"],
        "C": ["手动确认成本"],
        "D": ["尼日利亚工厂", "中国供应商", "尼日利亚本地供应商", "手动确认"],
        "E": ["尼日利亚工厂配套后出货", "中国配件直发客户", "中国配件发尼厂后配套", "尼日利亚本地采购配套"],
        "F": ["计入套装BOM", "计入独立物流", "客户单独报价", "仅参考不计入"],
        "G": ["是", "否"],
        "H": ["USD", "RMB"],
        "I": ["FOB尼日利亚港", "CFR目的港", "CIF目的港", "FOB中国港", "手动项目报价"],
        "J": ["套装总价", "分项报价"],
        "K": ["按柜", "按CBM", "按件", "按金额比例"],
        "L": ["启用", "停用"],
        "M": ["公式测算", "手动填写"],
        "N": ["手动分摊基数", "按装柜测算量", "按订单实际柜数摊销"],
    }
    for col, values in list_data.items():
        for i, value in enumerate(values, start=1):
            lists[f"{col}{i}"] = value
    add_name(wb, "NigeriaCostTerms", "Lists", "$A$1:$A$4")
    add_name(wb, "ChinaCostTerms", "Lists", "$B$1:$B$5")
    add_name(wb, "ManualCostTerms", "Lists", "$C$1:$C$1")
    add_name(wb, "QuoteSources", "Lists", "$D$1:$D$4")
    add_name(wb, "SupplyPaths", "Lists", "$E$1:$E$4")
    add_name(wb, "CostHandling", "Lists", "$F$1:$F$4")
    add_name(wb, "YesNo", "Lists", "$G$1:$G$2")
    add_name(wb, "Currencies", "Lists", "$H$1:$H$2")
    add_name(wb, "CustomerTerms", "Lists", "$I$1:$I$5")
    add_name(wb, "QuoteDisplayModes", "Lists", "$J$1:$J$2")
    add_name(wb, "AllocationMethods", "Lists", "$K$1:$K$4")
    add_name(wb, "EnableStatus", "Lists", "$L$1:$L$2")
    add_name(wb, "LogisticsInputModes", "Lists", "$M$1:$M$2")
    add_name(wb, "LogisticsBasisModes", "Lists", "$N$1:$N$3")

    # Fill sheet
    ws.merge_cells("A1:S1")
    ws["A1"] = "卫浴陶瓷报价表 - 第一版改造"
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A2"] = "销售只录入真实询价结果；表格按价格口径、供应链路径、成本处理方式自动判断BOM、物流、客户展示和风险。"
    ws.merge_cells("A2:S2")
    ws["A2"].alignment = left

    settings = [
        ("客户名称", "", "项目名称", "", "报价单号", "QT-20260529-001", "报价日期", "=TODAY()", "客户报价币种", "USD"),
        ("型号编号", "GD1A603-EL-S", "订单数量", 600, "报价汇率", 6.6, "目标成本加价率", 0.3, "底线成本加价率", 0),
        ("客户报价条款", "FOB尼日利亚港", "目的港", "Lagos-Apapa", "报价显示方式", "套装总价", "外购件成本加价率", 0.03, "是否DDP/项目例外", "否"),
        ("保险费率", 0, "保险金额RMB", 0, "手动项目报价审批人", "", "手动项目报价审批日期", "", "手动项目报价有效期", ""),
        ("尼日利亚默认装柜量", 600, "中国配件默认装柜量", 600, "手动项目报价原因", "", "客户备注", "", "内部备注", ""),
    ]
    for r, row in enumerate(settings, start=4):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c, value)
            ws.cell(r, c).alignment = left
            ws.cell(r, c).font = normal_font
        for c in [1, 3, 5, 7, 9]:
            ws.cell(r, c).fill = PatternFill("solid", fgColor=blue)
            ws.cell(r, c).font = Font(name="Microsoft YaHei", bold=True, size=10)
        for c in [2, 4, 6, 8, 10]:
            ws.cell(r, c).fill = PatternFill("solid", fgColor=light_yellow)
    ws["H5"].number_format = "0.00%"
    ws["J5"].number_format = "0.00%"
    ws["H6"].number_format = "0.00%"
    ws["B7"].number_format = "0.00%"
    ws["D7"].number_format = "#,##0.00"
    ws["H4"].number_format = "yyyy-mm-dd"

    image_placeholders = [
        ("K4:M8", "产品图片 1\n粘贴图片到此处"),
        ("N4:P8", "产品图片 2\n粘贴图片到此处"),
    ]
    for cell_range, label in image_placeholders:
        ws.merge_cells(cell_range)
        top_left = ws[cell_range.split(":")[0]]
        top_left.value = label
        top_left.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        top_left.font = Font(name="Microsoft YaHei", bold=True, size=11, color="666666")
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="F7F7F7")
                cell.border = border

    quick_review = [
        ("总成本RMB/件", "=InternalCost!B4+InternalCost!D4+InternalCost!B6+InternalCost!F4"),
        ("目标利润RMB/件", "=InternalCost!F7"),
        ("综合成本加价率", "=InternalCost!B10"),
        ("销售价USD/件", "=InternalCost!F9"),
        ("整单销售额USD", "=H9*$D$5"),
    ]
    for idx, (label, formula) in enumerate(quick_review):
        label_col = 1 + idx * 2
        value_col = label_col + 1
        ws.cell(9, label_col, label)
        ws.cell(9, value_col, formula)
        ws.cell(9, label_col).fill = PatternFill("solid", fgColor=blue)
        ws.cell(9, label_col).font = Font(name="Microsoft YaHei", bold=True, size=10)
        ws.cell(9, value_col).fill = PatternFill("solid", fgColor=gray)
        ws.cell(9, label_col).alignment = left
        ws.cell(9, value_col).alignment = left
        ws.cell(9, label_col).border = border
        ws.cell(9, value_col).border = border
    for cell in ["B9", "D9", "H9", "J9"]:
        ws[cell].number_format = "#,##0.00"
    ws["F9"].number_format = "0.00%"

    ws["A10"] = "成本录入区"
    ws.merge_cells("A10:P10")
    ws["A10"].fill = PatternFill("solid", fgColor=navy)
    ws["A10"].font = header_font
    ws["A10"].alignment = center
    headers = [
        "类别", "项目", "人民币金额", "报价来源", "价格口径", "供应链路径", "成本处理方式", "是否参与整体加价复核",
        "生效成本RMB", "手动原因", "审批人", "审批日期", "有效期", "备注", "物流取数方式", "物流公式测算RMB"
    ]
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(12, c, header)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    items = [
        ("VC cost陶瓷成本", "Tank/cistern lid水箱盖", None, "尼日利亚工厂", "尼厂出厂成本", "尼日利亚工厂配套后出货", "计入套装BOM", "是", "否", "", "", "", ""),
        ("VC cost陶瓷成本", "Tank/cistern 水箱", None, "尼日利亚工厂", "尼厂出厂成本", "尼日利亚工厂配套后出货", "计入套装BOM", "是", "否", "", "", "", ""),
        ("VC cost陶瓷成本", "toilet bowl/pan马桶", 47.8247810413976, "尼日利亚工厂", "尼厂出厂成本", "尼日利亚工厂配套后出货", "计入套装BOM", "是", "否", "", "", "", "旧表GD1A603-EL-S样例"),
        ("VC cost陶瓷成本", "basin/sink盆", None, "尼日利亚工厂", "尼厂出厂成本", "尼日利亚工厂配套后出货", "仅参考不计入", "否", "否", "", "", "", ""),
        ("VC cost陶瓷成本", "Pedestal", None, "尼日利亚工厂", "尼厂出厂成本", "尼日利亚工厂配套后出货", "仅参考不计入", "否", "否", "", "", "", ""),
        ("Accessories 配件", "Flush valve 水件", None, "中国供应商", "送中国指定仓/进仓价", "中国配件发尼厂后配套", "计入套装BOM", "是", "否", "", "", "", ""),
        ("Accessories 配件", "Toilet seat盖板", None, "中国供应商", "送中国指定仓/进仓价", "中国配件发尼厂后配套", "计入套装BOM", "是", "否", "", "", "", ""),
        ("Accessories 配件", "Overflow 溢水环", None, "中国供应商", "送中国指定仓/进仓价", "中国配件直发客户", "客户单独报价", "是", "否", "", "", "", ""),
        ("Accessories 配件", "Push button/lever按键、扳手", None, "中国供应商", "送中国指定仓/进仓价", "中国配件发尼厂后配套", "计入套装BOM", "是", "否", "", "", "", ""),
        ("Accessories 配件", "Tank fitting kit 水箱安装配件", None, "中国供应商", "送中国指定仓/进仓价", "中国配件发尼厂后配套", "计入套装BOM", "是", "否", "", "", "", ""),
        ("Accessories 配件", "Toilet fitting kit安装配件", None, "中国供应商", "送中国指定仓/进仓价", "中国配件发尼厂后配套", "计入套装BOM", "是", "否", "", "", "", ""),
        ("Accessories 配件", "others 其他配件", None, "中国供应商", "送中国指定仓/进仓价", "中国配件直发客户", "客户单独报价", "是", "否", "", "", "", ""),
        ("Package 包装", "Installation Manual说明书", 2, "中国供应商", "送中国指定仓/进仓价", "中国配件发尼厂后配套", "计入套装BOM", "是", "否", "", "", "", "旧表样例"),
        ("Package 包装", "Inner pack内包装", 6, "尼日利亚本地供应商", "尼厂装柜成本", "尼日利亚本地采购配套", "计入套装BOM", "是", "否", "", "", "", "旧表样例"),
        ("Package 包装", "outer carton外箱", 12.8, "尼日利亚本地供应商", "尼厂装柜成本", "尼日利亚本地采购配套", "计入套装BOM", "是", "否", "", "", "", "旧表样例"),
        ("Package 包装", "Pallet托盘", None, "尼日利亚本地供应商", "尼厂装柜成本", "尼日利亚本地采购配套", "仅参考不计入", "否", "否", "", "", "", ""),
        ("Labor Cost 人工成本", "Production labor cost生产人工成本", 6.24659047417853, "尼日利亚工厂", "尼厂出厂成本", "尼日利亚工厂配套后出货", "计入套装BOM", "是", "否", "", "", "", "旧表样例"),
        ("Labor Cost 人工成本", "Packing labor cost 包装人工成本", 0.97602987139373, "尼日利亚工厂", "尼厂装柜成本", "尼日利亚工厂配套后出货", "计入套装BOM", "是", "否", "", "", "", "旧表样例"),
        ("Spoilage 损耗", "Lost for low yield 合格率损耗", 9.43511900038598, "尼日利亚工厂", "手动确认成本", "尼日利亚工厂配套后出货", "计入套装BOM", "是", "否", "", "", "", "旧表样例"),
        ("Indirect Cost 间接费用", "Tooling 模具分摊", 2.55483100020165, "手动确认", "手动确认成本", "尼日利亚工厂配套后出货", "计入套装BOM", "是", "否", "", "", "", "旧表样例"),
        ("Indirect Cost 间接费用", "Management fee管理/财务/其他", 7, "尼日利亚工厂", "尼厂出厂成本", "尼日利亚工厂配套后出货", "计入套装BOM", "是", "否", "", "", "", "旧表样例"),
        ("Logistics 物流", "物流成本", None, "手动确认", "手动确认成本", "尼日利亚工厂配套后出货", "计入套装BOM", "否", "否", "", "", "", "选择公式测算时自动取计入BOM物流；选择手动填写时使用C列金额。物流不参与成本加价。"),
    ]
    start_row = 13
    end_row = start_row + len(items) - 1
    logistics_cost_row = end_row
    for idx, item in enumerate(items, start=start_row):
        values = list(item[:8]) + [None, "", "", "", "", item[12], "", ""]
        for c, value in enumerate(values, start=1):
            ws.cell(idx, c, value)
            ws.cell(idx, c).font = normal_font
            ws.cell(idx, c).alignment = left
            ws.cell(idx, c).border = border
        ws.cell(idx, 9, f'=N($C{idx})')
        for c in [3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 14, 15]:
            ws.cell(idx, c).fill = PatternFill("solid", fgColor=light_yellow)
        ws.cell(idx, 9).fill = PatternFill("solid", fgColor=gray)
        ws.cell(idx, 16).fill = PatternFill("solid", fgColor=gray)
        ws.cell(idx, 3).number_format = "#,##0.00"
        ws.cell(idx, 9).number_format = "#,##0.00"
        ws.cell(idx, 16).number_format = "#,##0.00"

    ws.cell(logistics_cost_row, 15, "公式测算")
    ws.cell(logistics_cost_row, 16, "=物流测算!C6")
    ws.cell(logistics_cost_row, 9, f'=IF($O{logistics_cost_row}="公式测算",$P{logistics_cost_row},$C{logistics_cost_row})')

    packaging_source_rows = [
        start_row + offset
        for offset, item in enumerate(items)
        if not (
            item[0].startswith("Labor Cost")
            or item[0].startswith("Spoilage")
            or item[0].startswith("Indirect Cost")
            or item[0].startswith("Package")
            or item[0].startswith("Logistics")
        )
    ]
    pack_home_title_row = end_row + 3
    pack_home_header_row = pack_home_title_row + 2
    pack_home_start_row = pack_home_header_row + 1
    pack_home_end_row = pack_home_start_row + len(packaging_source_rows) - 1
    ws.cell(pack_home_title_row, 1, "产品与包装信息区")
    ws.merge_cells(start_row=pack_home_title_row, start_column=1, end_row=pack_home_title_row, end_column=21)
    ws.cell(pack_home_title_row, 1).fill = PatternFill("solid", fgColor=navy)
    ws.cell(pack_home_title_row, 1).font = header_font
    ws.cell(pack_home_title_row, 1).alignment = center
    ws.cell(pack_home_title_row + 1, 1, "记录产品尺寸、重量、包装尺寸、毛重、每箱件数、CBM和装柜量；计入报价的项目未补齐时，检查页会报警。")
    ws.merge_cells(start_row=pack_home_title_row + 1, start_column=1, end_row=pack_home_title_row + 1, end_column=21)
    ws.cell(pack_home_title_row + 1, 1).alignment = left
    pack_headers = [
        "类别", "项目", "每套用量", "产品长mm", "产品宽mm", "产品高mm", "净重KG/件", "每箱件数",
        "包装长mm", "包装宽mm", "包装高mm", "毛重KG/箱", "整箱CBM", "单件CBM",
        "单件毛重KG", "体积可装件数", "重量可装件数", "建议装柜量", "限制因素", "供应链路径", "备注"
    ]
    for c, header in enumerate(pack_headers, start=1):
        cell = ws.cell(pack_home_header_row, c, header)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    for out_r, src_r in enumerate(packaging_source_rows, start=pack_home_start_row):
        ws.cell(out_r, 1, f"=A{src_r}")
        ws.cell(out_r, 2, f"=B{src_r}")
        ws.cell(out_r, 3, 1)
        ws.cell(out_r, 13, f'=IFERROR(IF(OR(I{out_r}="",J{out_r}="",K{out_r}=""),"",I{out_r}*J{out_r}*K{out_r}/1000000000),"")')
        ws.cell(out_r, 14, f'=IFERROR(IF(OR(M{out_r}="",H{out_r}=""),"",M{out_r}/MAX(H{out_r},1)),"")')
        ws.cell(out_r, 15, f'=IFERROR(IF(OR(L{out_r}="",H{out_r}=""),"",L{out_r}/MAX(H{out_r},1)),"")')
        ws.cell(out_r, 16, f'=IFERROR(IF(N{out_r}="","",ROUNDDOWN(装柜测算!$B$4*装柜测算!$B$6/MAX(N{out_r},0.0001),0)),"")')
        ws.cell(out_r, 17, f'=IFERROR(IF(O{out_r}="","",ROUNDDOWN(MAX(0,装柜测算!$B$5-装柜测算!$B$7)/MAX(O{out_r},0.0001),0)),"")')
        ws.cell(out_r, 18, f'=IFERROR(IF(OR(P{out_r}="",Q{out_r}=""),"",MIN(P{out_r},Q{out_r})),"")')
        ws.cell(out_r, 19, f'=IF(R{out_r}="","",IF(P{out_r}<Q{out_r},"体积限制",IF(Q{out_r}<P{out_r},"重量限制","体积/重量接近")))')
        ws.cell(out_r, 20, f"=F{src_r}")
        for c in range(1, 17):
            cell = ws.cell(out_r, c)
            cell.alignment = left
            cell.border = border
            cell.font = normal_font
        for c in range(17, 22):
            cell = ws.cell(out_r, c)
            cell.alignment = left
            cell.border = border
            cell.font = normal_font
        for c in range(3, 13):
            ws.cell(out_r, c).fill = PatternFill("solid", fgColor=light_yellow)
        ws.cell(out_r, 21).fill = PatternFill("solid", fgColor=light_yellow)
        for c in [13, 14, 15, 16, 17, 18, 19, 20]:
            ws.cell(out_r, c).fill = PatternFill("solid", fgColor=gray)
        for c in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
            ws.cell(out_r, c).number_format = "#,##0.00"
        for c in [16, 17, 18]:
            ws.cell(out_r, c).number_format = "#,##0"

    ws.freeze_panes = "A13"
    set_widths(ws, {
        "A": 18, "B": 30, "C": 13, "D": 16, "E": 20, "F": 22, "G": 16, "H": 16,
        "I": 14, "J": 18, "K": 12, "L": 12, "M": 12, "N": 18, "O": 22, "P": 24,
        "Q": 12, "R": 12, "S": 18, "T": 22, "U": 24
    })

    # Data validations on fill sheet
    add_list_validation(ws, "QuoteSources", [f"D{start_row}:D{end_row}"])
    for r in range(start_row, end_row + 1):
        add_list_validation(ws, f'=IF($D{r}="中国供应商",ChinaCostTerms,IF($D{r}="手动确认",ManualCostTerms,NigeriaCostTerms))', [f"E{r}"])
    add_list_validation(ws, "SupplyPaths", [f"F{start_row}:F{end_row}"])
    add_list_validation(ws, "CostHandling", [f"G{start_row}:G{end_row}"])
    add_list_validation(ws, "YesNo", [f"H{start_row}:H{end_row}"])
    add_list_validation(ws, "CustomerTerms", ["B6"])
    add_list_validation(ws, "QuoteDisplayModes", ["F6"])
    add_list_validation(ws, "Currencies", ["J4"])
    add_list_validation(ws, "YesNo", ["J6"])
    add_list_validation(ws, "LogisticsInputModes", [f"O{logistics_cost_row}"])

    # Product and packing information
    pack_info["A1"] = "产品与包装信息"
    pack_info.merge_cells("A1:U1")
    pack_info["A1"].fill = PatternFill("solid", fgColor=navy)
    pack_info["A1"].font = title_font
    pack_info["A1"].alignment = center
    pack_info["A2"] = "内部联动页：数据来自填写首页下方的产品与包装信息区，用于检查项和客户版包装信息展示。"
    pack_info.merge_cells("A2:U2")
    pack_info["A2"].alignment = left
    pack_headers = [
        "类别", "项目", "每套用量", "产品长mm", "产品宽mm", "产品高mm", "净重KG/件", "每箱件数",
        "包装长mm", "包装宽mm", "包装高mm", "毛重KG/箱", "整箱CBM", "单件CBM",
        "单件毛重KG", "体积可装件数", "重量可装件数", "建议装柜量", "限制因素", "供应链路径", "备注"
    ]
    for c, header in enumerate(pack_headers, start=1):
        cell = pack_info.cell(4, c, header)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    pack_info_start_row = 5
    pack_info_end_row = pack_info_start_row + len(packaging_source_rows) - 1
    for out_r, src_r in enumerate(packaging_source_rows, start=pack_info_start_row):
        home_r = pack_home_start_row + packaging_source_rows.index(src_r)
        pack_info.cell(out_r, 1, f"=填写首页!A{src_r}")
        pack_info.cell(out_r, 2, f"=填写首页!B{src_r}")
        for c in range(3, 19):
            col = get_column_letter(c)
            pack_info.cell(out_r, c, f'=IF(填写首页!{col}{home_r}="","",填写首页!{col}{home_r})')
        pack_info.cell(out_r, 19, f'=IF(填写首页!S{home_r}="","",填写首页!S{home_r})')
        pack_info.cell(out_r, 20, f"=填写首页!F{src_r}")
        pack_info.cell(out_r, 21, f'=IF(填写首页!U{home_r}="","",填写首页!U{home_r})')
        pack_info.cell(out_r, 22, f"=IFERROR(C{out_r}*N{out_r},0)")
        pack_info.cell(out_r, 23, f"=IFERROR(C{out_r}*O{out_r},0)")
        pack_info.cell(out_r, 24, f'=IF(OR(T{out_r}="尼日利亚工厂配套后出货",T{out_r}="中国配件发尼厂后配套",T{out_r}="尼日利亚本地采购配套"),1,0)')
        pack_info.cell(out_r, 25, f'=IF(T{out_r}="中国配件直发客户",1,0)')
        pack_info.cell(out_r, 26, f'=IF(B{out_r}<>"",1,0)')
        pack_info.cell(out_r, 27, f'=IF(B{out_r}="toilet bowl/pan马桶",1,0)')
        pack_info.cell(out_r, 28, f'=IF(AND(T{out_r}="中国配件发尼厂后配套",填写首页!$E${src_r}<>"到尼厂入库成本"),1,0)')
        for c in range(1, 29):
            cell = pack_info.cell(out_r, c)
            cell.alignment = left
            cell.border = border
            cell.font = normal_font
        for c in range(1, 29):
            pack_info.cell(out_r, c).fill = PatternFill("solid", fgColor=gray)
        for c in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
            pack_info.cell(out_r, c).number_format = "#,##0.00"
        for c in [16, 17, 18]:
            pack_info.cell(out_r, c).number_format = "#,##0"
    pack_info.freeze_panes = "A5"
    helper_headers = ["每套CBM_helper", "每套KG_helper", "尼日利亚组合_helper", "中国直发_helper", "全部实物_helper", "主产品_helper", "中国发尼厂_helper"]
    for c, header in enumerate(helper_headers, start=22):
        cell = pack_info.cell(4, c, header)
        cell.fill = PatternFill("solid", fgColor=dark_gray)
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    set_widths(pack_info, {
        "A": 18, "B": 30, "C": 11, "D": 11, "E": 11, "F": 11, "G": 11,
        "H": 11, "I": 11, "J": 11, "K": 12, "L": 12, "M": 12, "N": 12,
        "O": 12, "P": 14, "Q": 14, "R": 14, "S": 14, "T": 22, "U": 28
    })

    # Loading calculation
    load_calc["A1"] = "装柜测算"
    load_calc.merge_cells("A1:J1")
    load_calc["A1"].fill = PatternFill("solid", fgColor=navy)
    load_calc["A1"].font = title_font
    load_calc["A1"].alignment = center
    load_calc["A2"] = "按首页产品与包装信息估算：先算单项体积/重量限制，再按供应链路径组合成每套CBM和每套毛重，估算整柜可装套数。"
    load_calc.merge_cells("A2:J2")
    load_calc["A2"].alignment = left
    load_calc_params = [
        ("柜型", "40HQ"),
        ("有效容积CBM", 66),
        ("限重KG", 26500),
        ("装载效率", 0.92),
        ("安全重量余量KG", 1000),
    ]
    for idx, (label, value) in enumerate(load_calc_params, start=3):
        load_calc.cell(idx, 1, label)
        load_calc.cell(idx, 2, value)
        load_calc.cell(idx, 1).fill = PatternFill("solid", fgColor=blue)
        load_calc.cell(idx, 1).font = Font(name="Microsoft YaHei", bold=True, size=10)
        load_calc.cell(idx, 2).fill = PatternFill("solid", fgColor=light_yellow)
        load_calc.cell(idx, 1).border = border
        load_calc.cell(idx, 2).border = border
    load_calc["B6"].number_format = "0.00%"
    combo_header_row = 10
    combo_headers = ["组合场景", "每套CBM", "每套毛重KG", "体积可装套数", "重量可装套数", "建议装柜套数", "限制因素", "订单套数", "预计柜数", "说明"]
    for c, header in enumerate(combo_headers, start=1):
        cell = load_calc.cell(combo_header_row, c, header)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    scenario_rows = [
        ("尼日利亚整套出货组合", "X", "陶瓷、发尼厂后配套配件、本地采购配套一起按尼日利亚整套出货估算。"),
        ("中国配件直发组合", "Y", "只估算中国直发客户的配件组合，不混入尼日利亚整套出货。"),
        ("全部实物理论同柜", "Z", "仅供比较：假设所有实物同柜，不代表真实出货路径。"),
        ("主产品单项(toilet bowl/pan马桶)", "AA", "用于只有主产品单项出货时参考。"),
        ("中国配件发尼厂组合", "AB", "只估算中国发尼厂后配套的配件，用于中国到尼厂中转物流分摊。"),
    ]
    for r, (scenario, helper_col, note) in enumerate(scenario_rows, start=combo_header_row + 1):
        load_calc.cell(r, 1, scenario)
        load_calc.cell(r, 2, f"=SUMIFS(产品包装信息!$V${pack_info_start_row}:$V${pack_info_end_row},产品包装信息!${helper_col}${pack_info_start_row}:${helper_col}${pack_info_end_row},1)")
        load_calc.cell(r, 3, f"=SUMIFS(产品包装信息!$W${pack_info_start_row}:$W${pack_info_end_row},产品包装信息!${helper_col}${pack_info_start_row}:${helper_col}${pack_info_end_row},1)")
        load_calc.cell(r, 4, f'=IF(B{r}=0,"",ROUNDDOWN($B$4*$B$6/MAX(B{r},0.0001),0))')
        load_calc.cell(r, 5, f'=IF(C{r}=0,"",ROUNDDOWN(MAX(0,$B$5-$B$7)/MAX(C{r},0.0001),0))')
        load_calc.cell(r, 6, f'=IF(OR(D{r}="",E{r}=""),"",MIN(D{r},E{r}))')
        load_calc.cell(r, 7, f'=IF(F{r}="","",IF(D{r}<E{r},"体积限制",IF(E{r}<D{r},"重量限制","体积/重量接近")))')
        load_calc.cell(r, 8, "=填写首页!$D$5")
        load_calc.cell(r, 9, f'=IF(F{r}="","",ROUNDUP(H{r}/MAX(F{r},1),0))')
        load_calc.cell(r, 10, note)
        for c in range(1, 11):
            load_calc.cell(r, c).alignment = left
            load_calc.cell(r, c).border = border
            load_calc.cell(r, c).font = normal_font
        for c in range(2, 6):
            load_calc.cell(r, c).number_format = "#,##0.00"
        for c in [6, 8, 9]:
            load_calc.cell(r, c).number_format = "#,##0"

    detail_header_row = 18
    load_calc.cell(detail_header_row - 1, 1, "单项装柜明细")
    load_calc.merge_cells(start_row=detail_header_row - 1, start_column=1, end_row=detail_header_row - 1, end_column=10)
    load_calc.cell(detail_header_row - 1, 1).fill = PatternFill("solid", fgColor=navy)
    load_calc.cell(detail_header_row - 1, 1).font = header_font
    load_calc.cell(detail_header_row - 1, 1).alignment = center
    detail_headers = ["项目", "供应链路径", "每套用量", "单件CBM", "单件毛重KG", "体积可装件数", "重量可装件数", "建议装柜量", "限制因素", "备注"]
    for c, header in enumerate(detail_headers, start=1):
        cell = load_calc.cell(detail_header_row, c, header)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    for out_r, src_r in enumerate(range(pack_info_start_row, pack_info_end_row + 1), start=detail_header_row + 1):
        load_calc.cell(out_r, 1, f"=产品包装信息!B{src_r}")
        load_calc.cell(out_r, 2, f"=产品包装信息!T{src_r}")
        load_calc.cell(out_r, 3, f"=产品包装信息!C{src_r}")
        load_calc.cell(out_r, 4, f"=产品包装信息!N{src_r}")
        load_calc.cell(out_r, 5, f"=产品包装信息!O{src_r}")
        load_calc.cell(out_r, 6, f"=产品包装信息!P{src_r}")
        load_calc.cell(out_r, 7, f"=产品包装信息!Q{src_r}")
        load_calc.cell(out_r, 8, f"=产品包装信息!R{src_r}")
        load_calc.cell(out_r, 9, f"=产品包装信息!S{src_r}")
        load_calc.cell(out_r, 10, f"=产品包装信息!U{src_r}")
        for c in range(1, 11):
            load_calc.cell(out_r, c).alignment = left
            load_calc.cell(out_r, c).border = border
            load_calc.cell(out_r, c).font = normal_font
    set_widths(load_calc, {"A": 28, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 12, "I": 12, "J": 44})
    load_calc.freeze_panes = "A19"

    # Logistics params
    log_params.append(["费用名称", "金额RMB", "适用路径", "分摊方式", "分摊基数来源", "手动分摊基数", "生效分摊基数/装柜量", "预计柜数", "手动单件物流RMB", "单件摊销RMB", "是否启用", "备注"])
    params = [
        ["尼日利亚出货物流", 10000, "尼日利亚工厂配套后出货", "按柜", "按订单实际柜数摊销", "='填写首页'!$B$8", '=IF($E2="手动分摊基数",N($F2),IFERROR(IF(装柜测算!$F$11="",N($F2),装柜测算!$F$11),N($F2)))', '=IF(AND($D2="按柜",$E2="按订单实际柜数摊销"),ROUNDUP(填写首页!$D$5/MAX($G2,1),0),"")', "", '=IF($K2<>"是",0,IF(N($I2)>0,$I2,IF($D2="按柜",IF($E2="按订单实际柜数摊销",$B2*MAX($H2,1)/MAX(填写首页!$D$5,1),$B2/MAX($G2,1)),IF($D2="按件",$B2,IF($D2="按CBM",$B2*$G2,IF($D2="按金额比例",$B2*$G2,0))))))', "是", "默认按订单实际柜数摊销；包装信息未填时回退首页默认装柜量"],
        ["尼日利亚本地采购配套物流", 0, "尼日利亚本地采购配套", "按件", "手动分摊基数", 1, '=IF($E3="手动分摊基数",N($F3),IFERROR(IF(装柜测算!$F$11="",N($F3),装柜测算!$F$11),N($F3)))', '=IF(AND($D3="按柜",$E3="按订单实际柜数摊销"),ROUNDUP(填写首页!$D$5/MAX($G3,1),0),"")', "", '=IF($K3<>"是",0,IF(N($I3)>0,$I3,IF($D3="按柜",IF($E3="按订单实际柜数摊销",$B3*MAX($H3,1)/MAX(填写首页!$D$5,1),$B3/MAX($G3,1)),IF($D3="按件",$B3,IF($D3="按CBM",$B3*$G3,IF($D3="按金额比例",$B3*$G3,0))))))', "否", "本地采购配套默认不按出口整柜重复分摊；如有本地送货/提货费再填金额并启用"],
        ["中国配件直发物流", 11000, "中国配件直发客户", "按柜", "按订单实际柜数摊销", "='填写首页'!$D$8", '=IF($E4="手动分摊基数",N($F4),IFERROR(IF(装柜测算!$F$12="",N($F4),装柜测算!$F$12),N($F4)))', '=IF(AND($D4="按柜",$E4="按订单实际柜数摊销"),ROUNDUP(填写首页!$D$5/MAX($G4,1),0),"")', "", '=IF($K4<>"是",0,IF(N($I4)>0,$I4,IF($D4="按柜",IF($E4="按订单实际柜数摊销",$B4*MAX($H4,1)/MAX(填写首页!$D$5,1),$B4/MAX($G4,1)),IF($D4="按件",$B4,IF($D4="按CBM",$B4*$G4,IF($D4="按金额比例",$B4*$G4,0))))))', "是", "中国直发客户的默认物流参数；按中国直发组合装柜量"],
        ["中国到尼厂中转物流", 11000, "中国配件发尼厂后配套", "按柜", "按订单实际柜数摊销", "='填写首页'!$D$8", '=IF($E5="手动分摊基数",N($F5),IFERROR(IF(装柜测算!$F$15="",N($F5),装柜测算!$F$15),N($F5)))', '=IF(AND($D5="按柜",$E5="按订单实际柜数摊销"),ROUNDUP(填写首页!$D$5/MAX($G5,1),0),"")', "", '=IF($K5<>"是",0,IF(N($I5)>0,$I5,IF($D5="按柜",IF($E5="按订单实际柜数摊销",$B5*MAX($H5,1)/MAX(填写首页!$D$5,1),$B5/MAX($G5,1)),IF($D5="按件",$B5,IF($D5="按CBM",$B5*$G5,IF($D5="按金额比例",$B5*$G5,0))))))', "是", "仅当配件不是到尼厂入库成本时计入；按中国发尼厂组合装柜量"],
        ["尼厂到港/装柜/出口杂费", 0, "尼日利亚工厂配套后出货", "按柜", "按订单实际柜数摊销", "='填写首页'!$B$8", '=IF($E6="手动分摊基数",N($F6),IFERROR(IF(装柜测算!$F$11="",N($F6),装柜测算!$F$11),N($F6)))', '=IF(AND($D6="按柜",$E6="按订单实际柜数摊销"),ROUNDUP(填写首页!$D$5/MAX($G6,1),0),"")', "", '=IF($K6<>"是",0,IF(N($I6)>0,$I6,IF($D6="按柜",IF($E6="按订单实际柜数摊销",$B6*MAX($H6,1)/MAX(填写首页!$D$5,1),$B6/MAX($G6,1)),IF($D6="按件",$B6,IF($D6="按CBM",$B6*$G6,IF($D6="按金额比例",$B6*$G6,0))))))', "否", "尼厂出厂成本但需补本地费用时启用"],
    ]
    for row in params:
        log_params.append(row)
    for ws2 in [log_params]:
        for row in ws2.iter_rows():
            for cell in row:
                cell.alignment = left
                cell.border = border
                cell.font = normal_font
    style_range(ws2, "A1:L1", PatternFill("solid", fgColor=navy), header_font, center, border)
    for r in range(2, 7):
        for c in [5, 6, 9]:
            log_params.cell(r, c).fill = PatternFill("solid", fgColor=light_yellow)
        for c in [7, 8, 10]:
            log_params.cell(r, c).fill = PatternFill("solid", fgColor=gray)
    set_widths(log_params, {"A": 24, "B": 12, "C": 22, "D": 12, "E": 18, "F": 14, "G": 18, "H": 12, "I": 16, "J": 14, "K": 10, "L": 46})
    add_list_validation(log_params, "SupplyPaths", ["C2:C200"])
    add_list_validation(log_params, "AllocationMethods", ["D2:D200"])
    add_list_validation(log_params, "LogisticsBasisModes", ["E2:E200"])
    add_list_validation(log_params, "YesNo", ["K2:K200"])

    # Logistics calc
    log_calc.append(["物流路径", "是否触发", "计入BOM物流RMB/件", "独立物流RMB/件", "客户单独物流RMB/件", "说明"])
    active_cost_end_row = logistics_cost_row - 1
    logistics_rows = [
        ["尼日利亚工厂配套后出货", f'=SUMPRODUCT((填写首页!$F${start_row}:$F${active_cost_end_row}=A2)*(填写首页!$G${start_row}:$G${active_cost_end_row}<>"仅参考不计入")*(填写首页!$I${start_row}:$I${active_cost_end_row}>0))>0',
         '=IF(B2,SUMIFS(物流参数!$J:$J,物流参数!$C:$C,A2,物流参数!$K:$K,"是"),0)', 0, 0, "陶瓷/整套尼日利亚出货物流"],
        ["尼日利亚本地采购配套", f'=SUMPRODUCT((填写首页!$F${start_row}:$F${active_cost_end_row}=A3)*(填写首页!$G${start_row}:$G${active_cost_end_row}<>"仅参考不计入")*(填写首页!$I${start_row}:$I${active_cost_end_row}>0))>0',
         '=IF(B3,SUMIFS(物流参数!$J:$J,物流参数!$C:$C,A3,物流参数!$K:$K,"是"),0)', 0, 0, "本地采购随整套发货"],
        ["中国配件直发客户", f'=SUMPRODUCT((填写首页!$F${start_row}:$F${active_cost_end_row}=A4)*(填写首页!$G${start_row}:$G${active_cost_end_row}<>"仅参考不计入")*(填写首页!$I${start_row}:$I${active_cost_end_row}>0))>0',
         0,
         f'=IF(SUMPRODUCT((填写首页!$F${start_row}:$F${active_cost_end_row}=A4)*(填写首页!$G${start_row}:$G${active_cost_end_row}="计入独立物流")*(填写首页!$I${start_row}:$I${active_cost_end_row}>0))>0,SUMIFS(物流参数!$J:$J,物流参数!$C:$C,A4,物流参数!$K:$K,"是"),0)',
         f'=IF(SUMPRODUCT((填写首页!$F${start_row}:$F${active_cost_end_row}=A4)*(填写首页!$G${start_row}:$G${active_cost_end_row}="客户单独报价")*(填写首页!$I${start_row}:$I${active_cost_end_row}>0))>0,SUMIFS(物流参数!$J:$J,物流参数!$C:$C,A4,物流参数!$K:$K,"是"),0)',
         "中国直发物流，不混入尼日利亚整套物流"],
        ["中国配件发尼厂后配套", f'=SUMPRODUCT((填写首页!$F${start_row}:$F${active_cost_end_row}=A5)*(填写首页!$E${start_row}:$E${active_cost_end_row}<>"到尼厂入库成本")*(填写首页!$G${start_row}:$G${active_cost_end_row}<>"仅参考不计入")*(填写首页!$I${start_row}:$I${active_cost_end_row}>0))>0',
         '=IF(B5,SUMIFS(物流参数!$J:$J,物流参数!$C:$C,A5,物流参数!$K:$K,"是"),0)', 0, 0, "仅未包含中国到尼厂费用时补中转物流"],
    ]
    for row in logistics_rows:
        log_calc.append(row)
    log_calc.append(["合计", "", "=SUM(C2:C5)", "=SUM(D2:D5)", "=SUM(E2:E5)", ""])
    for row in log_calc.iter_rows():
        for cell in row:
            cell.alignment = left
            cell.border = border
            cell.font = normal_font
    style_range(log_calc, "A1:F1", PatternFill("solid", fgColor=navy), header_font, center, border)
    set_widths(log_calc, {"A": 24, "B": 12, "C": 18, "D": 18, "E": 18, "F": 34})

    # BOM
    bom.append(["类别", "项目", "生效成本RMB", "价格口径", "供应链路径", "成本处理方式", "计入BOM", "独立物流", "客户单独报价成本", "参与加价复核成本", "风险提示"])
    for out_r, src_r in enumerate(range(start_row, end_row + 1), start=2):
        bom.append([
            f"=填写首页!A{src_r}",
            f"=填写首页!B{src_r}",
            f"=填写首页!I{src_r}",
            f"=填写首页!E{src_r}",
            f"=填写首页!F{src_r}",
            f"=填写首页!G{src_r}",
            f'=IF(填写首页!G{src_r}="计入套装BOM",C{out_r},0)',
            f'=IF(填写首页!G{src_r}="计入独立物流",C{out_r},0)',
            f'=IF(填写首页!G{src_r}="客户单独报价",C{out_r},0)',
            f'=IF(填写首页!H{src_r}="是",C{out_r},0)',
            f'=IF(AND(D{out_r}="到尼厂入库成本",E{out_r}="中国配件发尼厂后配套"),"中国段已含，不再加中转物流",IF(AND(E{out_r}="中国配件直发客户",F{out_r}="计入套装BOM"),"路径冲突：直发客户不应进套装BOM",""))',
        ])
    material_total_r = end_row - start_row + 3
    bom.append(["成本小计", "", f"=SUM(C2:C{material_total_r-1})", "", "", "", f"=SUM(G2:G{material_total_r-1})", f"=SUM(H2:H{material_total_r-1})", f"=SUM(I2:I{material_total_r-1})", f"=SUM(J2:J{material_total_r-1})", ""])
    total_r = material_total_r + 1
    bom_logistics_r = logistics_cost_row - start_row + 2
    bom.append([
        "总计",
        "",
        f"=C{material_total_r}",
        "",
        "",
        "",
        f"=G{material_total_r}",
        f"=H{material_total_r}",
        f"=I{material_total_r}",
        f"=J{material_total_r}",
        "",
    ])
    for row in bom.iter_rows():
        for cell in row:
            cell.alignment = left
            cell.border = border
            cell.font = normal_font
    style_range(bom, "A1:K1", PatternFill("solid", fgColor=navy), header_font, center, border)
    set_widths(bom, {"A": 18, "B": 30, "C": 14, "D": 20, "E": 22, "F": 16, "G": 12, "H": 12, "I": 16, "J": 18, "K": 36})

    # InternalCost
    internal["A1"] = "内部复核版"
    internal.merge_cells("A1:F1")
    internal["A1"].fill = PatternFill("solid", fgColor=navy)
    internal["A1"].font = title_font
    internal["A1"].alignment = center
    internal_rows = [
        ("订单数量", "=填写首页!D5", "报价汇率", "=填写首页!F5", "客户条款", "=填写首页!B6"),
        ("套装BOM成本RMB/件", f"=BOM!G{total_r}", "客户单独报价物料成本RMB/件", f"=BOM!I{total_r}", "独立物流RMB/件", f"=BOM!H{total_r}+物流测算!D6"),
        ("套装物流成本RMB/件", f"=BOM!C{bom_logistics_r}", "套装加价基准成本RMB/件", "=MAX(B4-B5,0)", "目标成本加价率", "=填写首页!H5"),
        ("客户单独物流RMB/件", "=物流测算!E6", "客户单独加价基准成本RMB/件", "=D4", "外购件成本加价率", "=填写首页!H6"),
        ("套装目标利润RMB/件", "=D5*F5", "客户单独报价目标利润RMB/件", "=D6*F6", "整体目标利润RMB/件", "=B7+D7"),
        ("套装目标销售RMB/件", '=IFERROR(IF(填写首页!$B$6="CIF目的港",IF(填写首页!$D$7>0,D5+B5+B7+填写首页!$D$7/MAX(填写首页!$D$5,1),(D5+B5+B7)/(1-填写首页!$B$7)),D5+B5+B7),0)', "客户单独报价目标销售RMB/件", "=IFERROR(D6+B6+D7,0)", "整体目标销售RMB/件", "=B8+D8+F4"),
        ("套装目标销售USD/件", "=IFERROR(B8/填写首页!F5,0)", "客户单独报价USD/件", "=IFERROR(D8/填写首页!F5,0)", "整体目标销售USD/件", "=IFERROR(F8/填写首页!F5,0)"),
        ("当前综合成本加价率", "=IFERROR(F7/(D5+D6),0)", "底线成本加价率", "=填写首页!J5", "加价结论", '=IF(B10<D10,"低于底线","OK")'),
        ("CIF保险RMB/件", '=IF(填写首页!B6="CIF目的港",IF(填写首页!D7>0,填写首页!D7/MAX(填写首页!D5,1),B8*填写首页!B7),0)', "报价显示方式", "=填写首页!F6", "客户版可外发", '=IF(COUNTIF(价格口径检查!$B:$B,"报警")>0,"否","是")'),
    ]
    for r, row in enumerate(internal_rows, start=3):
        for c, val in enumerate(row, start=1):
            internal.cell(r, c, val)
            internal.cell(r, c).alignment = left
            internal.cell(r, c).border = border
            internal.cell(r, c).font = normal_font
        for c in [1, 3, 5]:
            internal.cell(r, c).fill = PatternFill("solid", fgColor=blue)
            internal.cell(r, c).font = Font(name="Microsoft YaHei", bold=True, size=10)
    set_widths(internal, {"A": 24, "B": 18, "C": 24, "D": 18, "E": 24, "F": 18})
    for cell in ["F5", "F6", "B10", "D10"]:
        internal[cell].number_format = "0.00%"
    for cell in ["B4", "D4", "F4", "B5", "D5", "B6", "D6", "B7", "D7", "F7", "B8", "D8", "F8", "B9", "D9", "F9", "B11"]:
        internal[cell].number_format = "#,##0.00"

    # CustomerQuote
    customer["A1"] = "Customer Quotation"
    customer.merge_cells("A1:H1")
    customer["A1"].fill = PatternFill("solid", fgColor=navy)
    customer["A1"].font = Font(name="Calibri", bold=True, size=16, color=white)
    customer["A1"].alignment = center
    customer_term_formula = '=IF(填写首页!B6="FOB尼日利亚港","FOB Nigeria Port",IF(填写首页!B6="CFR目的港","CFR Destination Port",IF(填写首页!B6="CIF目的港","CIF Destination Port",IF(填写首页!B6="FOB中国港","FOB China Port",IF(填写首页!B6="手动项目报价","Manual Project Quote",填写首页!B6)))))'
    display_mode_formula = '=IF(填写首页!F6="套装总价","Total Set Price",IF(填写首页!F6="分项报价","Itemized Price",填写首页!F6))'
    customer_rows = [
        ("Quote No.", "=填写首页!F4", "Date", "=填写首页!H4", "Customer", '=IF(填写首页!B4="","",填写首页!B4)', "Project", '=IF(填写首页!D4="","",填写首页!D4)'),
        ("Model", "=填写首页!B5", "Quantity", "=填写首页!D5", "Currency", "=填写首页!J4", "Trade Term", customer_term_formula),
        ("Destination Port", "=填写首页!D6", "Display Mode", display_mode_formula, "Validity", "", "Remark", '=IF(填写首页!H8="","",填写首页!H8)'),
    ]
    for r, row in enumerate(customer_rows, start=3):
        for c, val in enumerate(row, start=1):
            customer.cell(r, c, val)
            customer.cell(r, c).alignment = left
            customer.cell(r, c).border = border
            customer.cell(r, c).font = normal_font
        for c in [1, 3, 5, 7]:
            customer.cell(r, c).fill = PatternFill("solid", fgColor=blue)
            customer.cell(r, c).font = Font(name="Calibri", bold=True, size=10)
    customer["D3"].number_format = "yyyy-mm-dd"
    packing_customer = [
        "Packing Dimension",
        '=IFERROR(IF(OR(INDEX(产品包装信息!$I$5:$I$25,MATCH("toilet bowl/pan马桶",产品包装信息!$B$5:$B$25,0))="",INDEX(产品包装信息!$J$5:$J$25,MATCH("toilet bowl/pan马桶",产品包装信息!$B$5:$B$25,0))="",INDEX(产品包装信息!$K$5:$K$25,MATCH("toilet bowl/pan马桶",产品包装信息!$B$5:$B$25,0))=""),"",INDEX(产品包装信息!$I$5:$I$25,MATCH("toilet bowl/pan马桶",产品包装信息!$B$5:$B$25,0))&" x "&INDEX(产品包装信息!$J$5:$J$25,MATCH("toilet bowl/pan马桶",产品包装信息!$B$5:$B$25,0))&" x "&INDEX(产品包装信息!$K$5:$K$25,MATCH("toilet bowl/pan马桶",产品包装信息!$B$5:$B$25,0))&" mm"),"")',
        "Gross Weight",
        '=IFERROR(IF(INDEX(产品包装信息!$L$5:$L$25,MATCH("toilet bowl/pan马桶",产品包装信息!$B$5:$B$25,0))="","",INDEX(产品包装信息!$L$5:$L$25,MATCH("toilet bowl/pan马桶",产品包装信息!$B$5:$B$25,0))&" kg/carton"),"")',
        "PCS/Carton",
        '=IFERROR(IF(INDEX(产品包装信息!$H$5:$H$25,MATCH("toilet bowl/pan马桶",产品包装信息!$B$5:$B$25,0))="","",INDEX(产品包装信息!$H$5:$H$25,MATCH("toilet bowl/pan马桶",产品包装信息!$B$5:$B$25,0))),"")',
        "CBM/PCS",
        '=IFERROR(IF(INDEX(产品包装信息!$N$5:$N$25,MATCH("toilet bowl/pan马桶",产品包装信息!$B$5:$B$25,0))="","",INDEX(产品包装信息!$N$5:$N$25,MATCH("toilet bowl/pan马桶",产品包装信息!$B$5:$B$25,0))),"")',
    ]
    for c, val in enumerate(packing_customer, start=1):
        customer.cell(6, c, val)
        customer.cell(6, c).alignment = left
        customer.cell(6, c).border = border
        customer.cell(6, c).font = normal_font
    for c in [1, 3, 5, 7]:
        customer.cell(6, c).fill = PatternFill("solid", fgColor=blue)
        customer.cell(6, c).font = Font(name="Calibri", bold=True, size=10)
    quote_headers = ["Line", "Description", "Origin / Path", "Qty", "Unit Price", "Total Amount", "Trade Term", "Remark"]
    for c, h in enumerate(quote_headers, start=1):
        customer.cell(8, c, h)
        customer.cell(8, c).fill = PatternFill("solid", fgColor=navy)
        customer.cell(8, c).font = Font(name="Calibri", bold=True, color=white)
        customer.cell(8, c).alignment = center
        customer.cell(8, c).border = border
    customer_lines = [
        [1, '=IF(填写首页!F6="套装总价","Complete set total","Ceramic / complete set")', "Nigeria origin / set shipment", "=填写首页!D5", '=IF(填写首页!F6="套装总价",IF(填写首页!J4="RMB",InternalCost!F8,InternalCost!F9),IF(填写首页!J4="RMB",InternalCost!B8,InternalCost!B9))', '=D9*E9', customer_term_formula, '=IF(填写首页!F6="套装总价","Total display; internal BOM remains separated","")'],
        [2, '=IF(填写首页!F6="分项报价","China accessory separate quote","")', '=IF(填写首页!F6="分项报价","China origin / direct or separate path","")', '=IF(填写首页!F6="分项报价",填写首页!D5,"")', '=IF(填写首页!F6="分项报价",IF(填写首页!J4="RMB",InternalCost!D8,InternalCost!D9),"")', '=IF(填写首页!F6="分项报价",D10*E10,"")', '=IF(填写首页!F6="分项报价","FOB China Port / as agreed","")', ""],
    ]
    for r, line in enumerate(customer_lines, start=9):
        for c, val in enumerate(line, start=1):
            customer.cell(r, c, val)
            customer.cell(r, c).alignment = left
            customer.cell(r, c).border = border
            customer.cell(r, c).font = normal_font
    customer["A13"] = "Terms"
    customer["A14"] = "1. This customer version hides internal cost, supplier, margin and approval information."
    customer["A15"] = "2. CIF quotation requires insurance confirmation. DDP is not an automatic quotation term."
    customer["A16"] = "3. Final price is subject to product confirmation, valid quotation period, and shipping schedule."
    customer.merge_cells("A14:H14")
    customer.merge_cells("A15:H15")
    customer.merge_cells("A16:H16")
    set_widths(customer, {"A": 10, "B": 28, "C": 32, "D": 12, "E": 14, "F": 16, "G": 18, "H": 36})

    # Check sheet
    check["A1"] = "价格口径检查"
    check.merge_cells("A1:E1")
    check["A1"].fill = PatternFill("solid", fgColor=navy)
    check["A1"].font = title_font
    check["A1"].alignment = center
    check_headers = ["检查项", "状态", "处理建议", "关联位置", "说明"]
    for c, h in enumerate(check_headers, start=1):
        check.cell(3, c, h)
        check.cell(3, c).fill = PatternFill("solid", fgColor=navy)
        check.cell(3, c).font = header_font
        check.cell(3, c).alignment = center
        check.cell(3, c).border = border
    packaging_missing_conditions = []
    for home_r, src_r in zip(range(pack_home_start_row, pack_home_end_row + 1), packaging_source_rows):
        required_blank_check = ",".join(
            [f'填写首页!${col}${home_r}=""' for col in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]]
        )
        packaging_missing_conditions.append(
            f'AND(填写首页!$G${src_r}<>"仅参考不计入",填写首页!$I${src_r}>0,OR({required_blank_check}))'
        )
    packaging_missing_formula = f'=IF(OR({",".join(packaging_missing_conditions)}),"报警","OK")'

    check_rows = [
        ("价格金额为空", f'=IF(SUMPRODUCT((填写首页!$B${start_row}:$B${end_row}<>"")*(填写首页!$G${start_row}:$G${end_row}<>"仅参考不计入")*(填写首页!$O${start_row}:$O${end_row}<>"公式测算")*(填写首页!$C${start_row}:$C${end_row}=""))>0,"报警","OK")', "补齐人民币金额；物流成本选择公式测算时可不填C列", "填写首页 成本录入区", ""),
        ("价格口径为空", f'=IF(COUNTIFS(填写首页!$B${start_row}:$B${end_row},"<>",填写首页!$E${start_row}:$E${end_row},"")>0,"报警","OK")', "每个价格必须选择价格口径", "填写首页 E列", ""),
        ("动态口径不匹配", f'=IF(SUMPRODUCT(--(填写首页!$D${start_row}:$D${end_row}="中国供应商"),--ISNA(MATCH(填写首页!$E${start_row}:$E${end_row},Lists!$B$1:$B$5,0)))+SUMPRODUCT(--((填写首页!$D${start_row}:$D${end_row}="尼日利亚工厂")+(填写首页!$D${start_row}:$D${end_row}="尼日利亚本地供应商")>0),--ISNA(MATCH(填写首页!$E${start_row}:$E${end_row},Lists!$A$1:$A$4,0)))>0,"报警","OK")', "报价来源与价格口径必须匹配", "填写首页 D/E列", "中国供应商不能选择FOB尼日利亚港"),
        ("EXW价格但没有补物流", f'=IF(AND(COUNTIF(填写首页!$E${start_row}:$E${end_row},"EXW中国供应商")>0,SUMIFS(物流参数!$J:$J,物流参数!$C:$C,"中国配件直发客户",物流参数!$K:$K,"是")+SUMIFS(物流参数!$J:$J,物流参数!$C:$C,"中国配件发尼厂后配套",物流参数!$K:$K,"是")=0),"报警","OK")', "EXW只含供应商裸价，需要补物流或本地费用", "物流参数", ""),
        ("FOB/入库价格重复加本地费用", f'=IF(AND(COUNTIFS(填写首页!$E${start_row}:$E${end_row},"到尼厂入库成本",填写首页!$F${start_row}:$F${end_row},"中国配件发尼厂后配套")>0,填写首页!$O${logistics_cost_row}="手动填写",N(填写首页!$C${logistics_cost_row})>0),"报警","OK")', "到尼厂入库成本已含中国段；公式测算会自动剔除中转物流，手动物流需确认未重复包含中国到尼厂费用", "填写首页 E/F列 + 物流成本行", ""),
        ("中国直发配件被计入尼日利亚物流", f'=IF(COUNTIFS(填写首页!$F${start_row}:$F${end_row},"中国配件直发客户",填写首页!$G${start_row}:$G${end_row},"计入套装BOM")>0,"报警","OK")', "中国直发客户应客户单独报价或计入独立物流，不能混入套装BOM", "填写首页 F/G列", ""),
        ("CIF缺保险", '=IF(AND(填写首页!$B$6="CIF目的港",N(填写首页!$B$7)<=0,N(填写首页!$D$7)<=0),"报警","OK")', "CIF目的港必须填写保险费率或保险金额", "填写首页 B7/D7", ""),
        ("CFR/CIF没有目的港", '=IF(AND(OR(填写首页!$B$6="CFR目的港",填写首页!$B$6="CIF目的港"),填写首页!$D$6=""),"报警","OK")', "CFR/CIF必须有目的港", "填写首页 D6", ""),
        ("手动确认成本未审批", f'=IF(SUMPRODUCT((填写首页!$E${start_row}:$E${end_row}="手动确认成本")*(填写首页!$O${start_row}:$O${end_row}<>"公式测算")*(((填写首页!$J${start_row}:$J${end_row}="")+(填写首页!$K${start_row}:$K${end_row}="")+(填写首页!$L${start_row}:$L${end_row}="")+(填写首页!$M${start_row}:$M${end_row}=""))>0))>0,"报警","OK")', "内部成本例外必须填写原因、审批人、日期、有效期；物流公式测算不按手动成本审批", "填写首页 J:M列", ""),
        ("手动项目报价未审批", '=IF(AND(填写首页!$B$6="手动项目报价",OR(填写首页!$F$7="",填写首页!$H$7="",填写首页!$J$7="",填写首页!$F$8="")),"报警","OK")', "客户销售报价例外必须单独审批", "填写首页 F7:J7 / F8", ""),
        ("产品/包装信息缺失", packaging_missing_formula, "计入报价的实物项目需要补产品尺寸、净重、每箱件数、包装尺寸和毛重", "填写首页 产品与包装信息区", ""),
        ("成本加价率低于底线", '=IF(InternalCost!$B$10<InternalCost!$D$10,"报警","OK")', "复核目标价或申请审批", "InternalCost", ""),
        ("DDP被当作自动报价使用", '=IF(填写首页!$J$6="是","报警","OK")', "DDP只能作为手动项目报价并主管审批", "填写首页 J6", ""),
        ("客户版可外发", '=IF(COUNTIF($B$4:$B$16,"报警")>0,"否","是")', "否时不要发送CustomerQuote", "CustomerQuote", ""),
    ]
    for r, row in enumerate(check_rows, start=4):
        for c, val in enumerate(row, start=1):
            check.cell(r, c, val)
            check.cell(r, c).alignment = left
            check.cell(r, c).border = border
            check.cell(r, c).font = normal_font
    set_widths(check, {"A": 24, "B": 12, "C": 46, "D": 28, "E": 34})
    for r in range(4, 4 + len(check_rows)):
        check.cell(r, 2).fill = PatternFill("solid", fgColor=light_green)

    # General formatting
    for sheet in [ws, pack_info, load_calc, check, bom, log_params, log_calc, customer, internal, lists]:
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and cell.font == Font():
                    cell.font = normal_font

    # Hide support list sheet and protect formula-heavy sheets lightly.
    lists.sheet_state = "hidden"
    pack_info.sheet_state = "hidden"
    for protected in [pack_info, check, bom, log_calc, customer, internal]:
        protected.protection.sheet = True
        protected.protection.password = "quote"

    # Conditional formatting cannot be relied on for all clients, but status colors help after calculation.
    # Apply number formats.
    for sheet in [bom, log_params, log_calc, internal, customer]:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.number_format = "#,##0.00"
    customer["D3"].number_format = "yyyy-mm-dd"

    try:
        wb.save(OUTPUT)
        print(OUTPUT)
    except PermissionError:
        fallback = ROOT / "卫浴陶瓷报价表_第一版改造_20260529_物流BOM明细版.xlsx"
        wb.save(fallback)
        print(fallback)


if __name__ == "__main__":
    build()
